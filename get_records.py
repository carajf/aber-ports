### PYTHON SCRIPT TO CONVERT CSV (Excel) FILES INTO JSON 

import sys
from openpyxl import load_workbook
import json

# ship_excel_schema contains facts about about how the ship data is organized in Excel
import ship_excel_schema

# Function get_shipreturns a list of JSON ship objects
def get_records(excel_ship_file, u_key):
    current_key = u_key

    # ships is the list of JSON ship objects that will be returned
    records = []

    # load the Excel workbook
    try:

        wb = load_workbook(filename = excel_ship_file)

        # getting the name, number, and port registry of each ship (these don't change per worksheet nor per mariner record)
        vessel_name = wb.active[ship_excel_schema.vessel_name].value
        official_number = wb.active[ship_excel_schema.official_number].value
        port_of_registry = wb.active[ship_excel_schema.port_of_registry].value

        # get the start row for mariner data
        mariners_start_row = ship_excel_schema.mariners_start_row
        
        # for each sheet in the workbook, assign the value to a key in a ship dictionary
        for sheet in wb:

            # giving the start row and column
            row = mariners_start_row
            col = 1

            # whilst the cell in column x and row x is not empty
            while not (sheet.cell(column=col, row=row).value is None):
                
                # Create a new record
                record = {}
                
                # Assign each new record/row a unique key
                record["_id"] = u_key

                # Populate the record's ship data
                record["vessel name"] = vessel_name
                record["official number"] = official_number
                record["port of registry"] = port_of_registry

                # Populate the record's mariner data
                for attr, col in ship_excel_schema.mariner_attributes.items():
                    if not sheet.cell(column=col, row=row).value is None:

                        # Assigning mariner data
                        record[attr] = sheet.cell(column=col, row=row).value

                # Append the record to the records list
                records.append(record)

                # Next row and next unique key
                u_key += 1 
                row += 1
                col = 1

    except:
        # deal with the error if the file cannot be opened
        print(excel_ship_file)

    # Output of the function returns the records list of records
    return records, u_key

    
   
if __name__ == "__main__":
    # get the name of the Excel File
    print("Input the name of the Excel Ship File")
    excel_ship_file = sys.stdin.readline().strip()
    
    records = get_records(excel_ship_file)
    for record in records:
        print(json.dumps(record))
            
