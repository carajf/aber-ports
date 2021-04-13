### PYTHON SCRIPT TO CONVERT CSV (Excel) FILES INTO JSON 

# Import necessary modules (including get_records script which retrieves the data from the xlxs files and converts to JSON)
import os
import sys
from openpyxl import load_workbook
import json

# ship_excel_schema contains facts about about how the ship data is organized in Excel
import ship_excel_schema


def count_records(excel_ship_file):

  this_file_count = 0

  # load the Excel workbook
  try:
    wb = load_workbook(filename = excel_ship_file)
    
    # get the start row for mariner data
    mariners_start_row = ship_excel_schema.mariners_start_row
    
    # for each sheet in the workbook...
    for sheet in wb:

      # giving the start row and column
      start_row = mariners_start_row
      col = 1

      # whilst the cell in column x and row x is not empty
      while not (sheet.cell(column=col, row=start_row).value is None):
        this_file_count += 1
        start_row += 1

  except:
    # deal with the error if the file cannot be opened
    print(excel_ship_file)

  # Output of the function returns 
  return this_file_count




init_count = 0

def run_count_records(init_count):

  record_count = init_count

  # Locating the root directory of all ship data
  ships_dir = 'test\ABERSHIP_transcription_vtls004566921'

  # Walking through each file in the directory
  for root, dirs, files in os.walk(ships_dir):
    for file in files:
      name, ext = os.path.splitext(file)
      if ext == '.xlsx':
          
          record_count += count_records(os.path.join(root, file))

  print("The number of excel mariner records in all worksheets and workbooks is: ", record_count)

  




def count_documents():
  # Establish connection to mongodb.
  import pymongo
  from pymongo import MongoClient

  client = MongoClient('mongodb://localhost:27017')
  db = client.aber_ship_db
  coll = db.test_new_keys
  result = coll.count_documents({})

  print("The number of documents in collection: ", result)
   


run_count_records(init_count)
count_documents()


