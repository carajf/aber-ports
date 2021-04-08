### PYTHON SCRIPT TO CONVERT CSV (Excel) FILES INTO JSON 

import sys
from openpyxl import load_workbook
import json

# ship_excel_schema contains facts about about how the ship data is organized in Excel
import ship_excel_schema

# Function get_shipreturns a list of JSON ship objects
def get_ships(excel_ship_file):

    # ships is the list of JSON ship objects that will be returned
    ships = []

    # load the Excel workbook
    try:
        wb = load_workbook(filename = excel_ship_file)

        # getting the name, number, and port registry of each ship
        vessel_name = wb.active[ship_excel_schema.vessel_name].value
        official_number = wb.active[ship_excel_schema.official_number].value
        port_of_registry = wb.active[ship_excel_schema.port_of_registry].value
        
        # for each sheet in the workbook, assign the value to a key in a ship dictionary
        for sheet in wb:
            # get ship attributes from the first worksheet in the workbook
            ship = {}
            ship["vessel name"] = vessel_name
            ship["official number"] = official_number
            ship["port of registry"] = port_of_registry
            ship["mariners"] = []

            # get the start row for mariner data
            mariners_start_row = ship_excel_schema.mariners_start_row

            # iterate over the worksheets to find all the mariner data
            row = mariners_start_row
            col = 1
            while not (sheet.cell(column=col, row=row).value is None):
                mariner = {}
                for attr, col in ship_excel_schema.mariner_attributes.items():
                    if not sheet.cell(column=col, row=row).value is None:
                        mariner[attr] = sheet.cell(column=col, row=row).value
                ship["mariners"].append(mariner)
                row += 1

            ships.append(ship)

    except:
        # deal with the error if the file cannot be opened
        print(excel_ship_file)

    # Output of the function returns the ships list of dictionaries
    return ships
   
if __name__ == "__main__":
    # get the name of the Excel File
    print("Input the name of the Excel Ship File")
    excel_ship_file = sys.stdin.readline().strip()
    
    ships = get_ships(excel_ship_file)
    for ship in ships:
        print(json.dumps(ship))
            
